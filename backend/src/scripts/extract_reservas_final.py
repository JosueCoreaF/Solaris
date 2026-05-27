"""
Extractor de Reservas Excel → Supabase  (v3 FINAL)
====================================================
Lógica de colores DEFINITIVA (confirmada visualmente):

  BG=theme:0,  FONT=NONE o FF000000        → PAGADA        (texto negro, sin relleno)
  BG=theme:0,  FONT=theme:1                → PAGADA        (texto negro tema, sin relleno)
  BG=theme:5 (naranja/café), FONT=FF00B0F0 → NOCHE PENDIENTE POR CONSUMIR  (crédito)
  BG=FFFFFF00, FONT=FFFF0000, bold         → CORTESÍA/CANJE
  BG=FFFFC000, FONT=NONE                   → NOCHE PENDIENTE (naranja sólido)
  BG=FF00B050, FONT=FFFF0000               → NO DISPONIBLE (verde)
  BG=FFFF0000, FONT=FF00B0F0               → NO DISPONIBLE (rojo)
  * DISPONIBLE / texto rojo sin reserva    → OMITIDO

  Estado reserva (check_in):
    → Fondo amarillo FFFFFF00 = ya en hotel (check_in)
    → Resto = confirmada
"""

import openpyxl, re, json, sys, os, tempfile
from datetime import date, timedelta
from collections import defaultdict

ROOM_MAP = {
    5:  {"numero": 3,  "nombre": "3"},
    6:  {"numero": 4,  "nombre": "4"},
    7:  {"numero": 5,  "nombre": "5"},
    8:  {"numero": 6,  "nombre": "6"},
    9:  {"numero": 7,  "nombre": "7"},
    10: {"numero": 8,  "nombre": "8"},
    11: {"numero": 9,  "nombre": "9"},
    12: {"numero": 10, "nombre": "10"},
    13: {"numero": 11, "nombre": "11"},
    14: {"numero": 12, "nombre": "12"},
    15: {"numero": 13, "nombre": "13"},
    16: {"numero": 14, "nombre": "14"},
}

MES_NUM = {
    'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
    'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12
}

def get_bg(cell):
    try:
        f = cell.fill.fgColor
        if f.type == 'rgb':   return ('rgb', f.rgb)
        if f.type == 'theme': return ('theme', f.theme, f.tint)
    except: pass
    return ('none',)

def get_font_color(cell):
    try:
        c = cell.font.color
        if c.type == 'rgb':   return ('rgb', c.rgb)
        if c.type == 'theme': return ('theme', c.theme)
    except: pass
    return ('none',)

def classify(cell):
    """Retorna (tipo, estado_pago, estado_reserva) o None para omitir."""
    val = str(cell.value).strip() if cell.value else ""
    if not val or val in ['.', '0']:
        return None
    if val.startswith('* DISPONIBLE') or val == 'DISPONIBLE':
        return None  # omitir disponibles

    bg   = get_bg(cell)
    font = get_font_color(cell)
    bold = cell.font.bold if cell.font else False

    # ── NO DISPONIBLE ────────────────────────────────────────
    if bg == ('rgb', 'FF00B050'):           # verde + rojo texto
        return ('no_disponible', None, None)
    if bg == ('rgb', 'FFFF0000'):           # rojo + azul texto
        return ('no_disponible', None, None)

    # ── CORTESÍA (canje) – amarillo fondo + rojo texto + bold ─
    if bg[0] == 'rgb' and bg[1] == 'FFFFFF00' and font == ('rgb','FFFF0000') and bold:
        return ('reserva', 'cortesia', 'check_in')


    # ── NOCHE PENDIENTE theme:5 (naranja/café) + texto azul ────────
    if bg[0] == 'theme' and bg[1] == 5 and font == ('rgb','FF00B0F0'):
        return ('reserva', 'credito', 'confirmada')  # noche pendiente por consumir (crédito)

    # ── CORTESÍA – texto azul sin relleno (crédito/cortesía) ─
    if bg[0] in ('none','theme') and font == ('rgb','FF00B0F0'):
        return ('reserva', 'credito', 'confirmada')

    # ── PAGADA – texto negro sin relleno especial ─────────────
    if bg[0] == 'theme' and font[0] in ('none', 'rgb') and (font == ('none',) or font[1] == 'FF000000'):
        return ('reserva', 'pagado', 'check_in' if bold else 'confirmada')

    # ── PAGADA – texto negro tema ─────────────────────────────
    if bg[0] == 'theme' and font[0] == 'theme':
        return ('reserva', 'pagado', 'check_in' if bold else 'confirmada')

    # fallback
    return ('reserva', 'pagado', 'confirmada')


def parse_reserva(text, estado_pago=None):
    r = {
        "empresa": None, "huesped_principal": None,
        "huespedes_adicionales": [], "quien_reservo": None,
        "num_habitaciones": 1, "duracion_tipo": "noche", "duracion_valor": 1,
        "total_reserva": None, "telefono": None, "factura": None,
        "canal_reserva": "directo", "es_cortesia": False,
        "observaciones": None, "texto_raw": text,
    }
    if not text: return r
    t = text.upper()

    # Canal
    if 'EXPEDIA' in t:   r["canal_reserva"] = "expedia"
    elif 'AIRBNB' in t:  r["canal_reserva"] = "airbnb"
    elif 'WHATSAPP' in t: r["canal_reserva"] = "whatsapp"
    elif 'CANJE' in t:   r["canal_reserva"] = "canje"; r["es_cortesia"] = True

    # Precio
    m = re.search(r'L[.\s]*([\d,]+(?:\.\d+)?)', text, re.I)
    if m:
        try: r["total_reserva"] = float(m.group(1).replace(',',''))
        except: pass

    # Duración
    m = re.search(r'(\d+)\s*HORA', t)
    if m: r["duracion_tipo"] = "horas"; r["duracion_valor"] = int(m.group(1))
    else:
        m = re.search(r'(\d+)\s*NOCHE', t)
        if m: r["duracion_valor"] = int(m.group(1))

    # Teléfono
    m = re.search(r'CEL[.\s]*[+]?([\d\s\-]+)', text, re.I)
    if m: r["telefono"] = m.group(1).strip().rstrip('.')

    # Factura
    m = re.search(r'FACT(?:URA)?[#\s]*(0*\d+)', text, re.I)
    if m: r["factura"] = m.group(1)

    # Quien reservó
    m = re.search(r'TONITO LE RESERVO', text, re.I)
    if m: r["quien_reservo"] = "Tonito"
    else:
        m = re.search(r'(.+?)\s+LE RESERVO', text, re.I)
        if m: r["quien_reservo"] = m.group(1).strip().title()

    # Empresa y huéspedes
    EMPRESAS = ['EMISORAS UNIDAS','EXPEDIA','AIRBNB','TECUN','KRISTA MARISSA',
        'IPSA','ALPHA SERVICIO MEDICO','FINCA SA','MEYKO','CASM','ORO MAYA',
        'ORO MAYA DE HONDURAS','BIMBO','ODEF','FEFASA','SOLCOB',
        'CONFEDERACION DE TRABAJADORES DE HONDURAS','DISTRIBUIDORA EDT',
        'RADIO NORTE','POS DE HONDURAS',]

    clean = re.sub(r'^RESERVA\s+', '', text.strip(), flags=re.I)
    clean = re.sub(r'\s+\d+\s*HABITACION.*', '', clean, flags=re.I)
    clean = re.sub(r'\s+UNA\s+HABITACION.*', '', clean, flags=re.I)
    parts = [p.strip() for p in clean.split('/') if p.strip()]

    def clean_name(n):
        # Remove common noise tokens at the end
        n = re.sub(r'\s*(?:UNA|1)?\s*HABI[T|A]+CION.*', '', n, flags=re.I)
        n = re.sub(r'\s*POR\s+\d+\s*(?:HORAS?|NOCHES?).*', '', n, flags=re.I)
        n = re.sub(r'\s*A\s*L\.?\s*[\d,.]+.*', '', n, flags=re.I)
        n = re.sub(r'\s*CEL\.?\s*[\d-]+.*', '', n, flags=re.I)
        n = re.sub(r'\s*FACT(?:URA)?.*', '', n, flags=re.I)
        return n.strip().upper()

    if len(parts) >= 2 and estado_pago == 'credito':
        p0 = parts[0].upper().strip()
        es_emp = any(p0.startswith(e) for e in EMPRESAS)
        if es_emp:
            r["empresa"] = parts[0].strip().title()
            huespedes = [clean_name(p)
                         for p in parts[1:]
                         if p.strip() and not re.match(r'^[A-Z\s]+LE\s+RESERVO',p,re.I)]
            huespedes = [h for h in huespedes if h and len(h) > 2]
            if huespedes:
                r["huesped_principal"] = huespedes[0]
                r["huespedes_adicionales"] = huespedes[1:]
        else:
            huespedes = [clean_name(p)
                         for p in parts
                         if p.strip() and not re.match(r'^[A-Z\s]+LE\s+RESERVO',p,re.I)]
            huespedes = [h for h in huespedes if h and len(h) > 2]
            if huespedes:
                r["huesped_principal"] = huespedes[0]
                r["huespedes_adicionales"] = huespedes[1:]
    elif parts:
        nombre = clean_name(parts[0])
        r["huesped_principal"] = nombre

    # Observaciones
    obs = []
    if 'DESCUENTO' in t: obs.append('Descuento aplicado')
    if 'TERCERA EDAD' in t: obs.append('Descuento tercera edad')
    if 'NO QUIERE LIMPIEZA' in t: obs.append('No desea limpieza')
    if 'PAGANDO EN EFECTIVO' in t: obs.append('Pago en efectivo')
    m = re.search(r'(\d+)\s*PERSONAS', text, re.I)
    if m: obs.append(f"{m.group(1)} personas")
    m = re.search(r'SALE A LAS\s*([\d:.]+\s*[APM]*)', text, re.I)
    if m: obs.append(f"Sale a las {m.group(1)}")
    # Noche pendiente de consumo
    if 'NO LA CONSUMIO' in t or 'PENDIENTE DE CONSUMO' in t or 'DEJA PENDIENTE' in t:
        obs.append('Noche pendiente de consumo')
    if obs: r["observaciones"] = "; ".join(obs)

    return r


def extract(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'Hoja1' in wb.sheetnames:
        ws = wb['Hoja1']
    else:
        ws = wb.active

    mes = str(ws['A1'].value or '').strip()
    anio = ws['A2'].value
    mes_num = MES_NUM.get(mes.upper(), 1)

    registros = []
    stats = defaultdict(int)
    active_res = {col: None for col in ROOM_MAP}

    for row in range(1, ws.max_row + 1):
        dia_num_raw = ws.cell(row, 1).value
        
        # Check if this row is a Month header
        if isinstance(dia_num_raw, str) and dia_num_raw.strip().upper() in MES_NUM:
            mes = dia_num_raw.strip().upper()
            mes_num = MES_NUM.get(mes, 1)
            anio_raw = ws.cell(row + 1, 1).value
            if anio_raw:
                try: anio = int(anio_raw)
                except: pass
            continue
            
        # Check if this row is a Year header or empty
        if not dia_num_raw or str(dia_num_raw).strip() == str(anio):
            continue

        try: 
            dia_num = int(dia_num_raw)
            fecha = date(int(anio), mes_num, dia_num)
        except: 
            continue # If we can't parse the day number as integer, skip the row

        dia_sem = ws.cell(row, 2).value
        recep   = ws.cell(row, 3).value
        evento  = ws.cell(row, 4).value

        for col, room in ROOM_MAP.items():
            cell = ws.cell(row, col)
            raw  = str(cell.value).strip() if cell.value else ""
            if not raw or raw in ['.','0'] or raw.startswith('* DISPONIBLE') or raw == 'DISPONIBLE': 
                if active_res[col]:
                    registros.append(active_res[col])
                    active_res[col] = None
                continue

            clf = classify(cell)
            if clf is None:
                if active_res[col]:
                    registros.append(active_res[col])
                    active_res[col] = None
                stats['omitido'] += 1
                continue

            tipo, estado_pago, estado_reserva = clf

            from datetime import datetime
            today = datetime.now()

            if tipo == 'no_disponible':
                if not re.search(r'(?i)RESERVA\b', raw):
                    if active_res[col]:
                        registros.append(active_res[col])
                        active_res[col] = None
                    stats['no_disponible'] += 1
                    continue
                else:
                    tipo = 'reserva'
                    if estado_pago is None:
                        # ci aún no está definido aquí, lo derivamos de fecha
                        ci_temp = fecha if fecha else None
                        if ci_temp and ci_temp < today.date():
                            estado_pago = 'pagado'
                            estado_reserva = 'check_out'
                        else:
                            estado_pago = 'deuda'
                            estado_reserva = 'confirmada'

            stats[f"pago:{estado_pago}"] += 1

            # Divide si hay múltiples reservas separadas por "/ RESERVA" o "/RESERVA"
            raw_parts = re.split(r'(?i)/(?=\s*RESERVA\b)', raw)
            for part in raw_parts:
                part = part.strip()
                if not part: continue
                if re.match(r'(?i)^(SUCIA|MANTENIMIENTO|NO\s+DISPONIBLE)\b', part):
                    continue

                p = parse_reserva(part, estado_pago)

                # check_in / check_out
                dur_tipo = p["duracion_tipo"]
                dur_val  = p["duracion_valor"]
                if fecha:
                    ci = fecha
                    if dur_tipo == "horas":
                        co = fecha
                        tipo_res = "hora"
                    else:
                        co = fecha + timedelta(days=dur_val)
                        tipo_res = "noche"
                else:
                    ci = co = None
                    tipo_res = dur_tipo

                obs_parts = []
                if p["observaciones"]: obs_parts.append(p["observaciones"])
                if p["canal_reserva"] not in ("directo","canje"): obs_parts.append(f"Canal: {p['canal_reserva']}")
                if recep: obs_parts.append(f"Recepcionista: {recep}")
                obs_parts.append(f"[EXCEL:{part[:80]}]")

                new_res = {
                    "fecha":              ci.isoformat() if ci else None,
                    "check_out":          co.isoformat() if co else None,
                    "mes":                mes,
                    "anio":               int(anio) if anio else None,
                    "dia_numero":         int(dia_num),
                    "dia_semana":         str(dia_sem) if dia_sem else None,
                    "habitacion_numero":  room["numero"],
                    "habitacion_nombre":  room["nombre"],
                    "habitacion_alias":   room.get("alias", ""),
                    "estado_pago":        estado_pago,
                    "estado_reserva":     estado_reserva,
                    "tipo_reserva":       tipo_res,
                    "empresa":            p["empresa"],
                    "huesped_principal":  p["huesped_principal"],
                    "huespedes_adicionales": p["huespedes_adicionales"],
                    "quien_reservo":      p["quien_reservo"],
                    "total_reserva":      p["total_reserva"],
                    "telefono":           p["telefono"],
                    "factura":            p["factura"],
                    "canal_reserva":      p["canal_reserva"],
                    "es_cortesia":        p["es_cortesia"],
                    "observaciones":      "; ".join(obs_parts),
                    "texto_raw":          part,
                }

                if active_res[col]:
                    ar = active_res[col]
                    is_same = False
                    if p["huesped_principal"] and ar["huesped_principal"] == p["huesped_principal"]:
                        is_same = True
                    elif not p["huesped_principal"] and not ar["huesped_principal"] and part == ar["texto_raw"]:
                        is_same = True
                    
                    if is_same and ar["tipo_reserva"] != "hora":
                        if co: ar["check_out"] = co.isoformat()
                        if p["total_reserva"]:
                            ar["total_reserva"] = (ar["total_reserva"] or 0) + p["total_reserva"]
                        if p["observaciones"] and p["observaciones"] not in ar["observaciones"]:
                            ar["observaciones"] += "; " + p["observaciones"]
                        if estado_pago == 'pagado':
                            ar["estado_pago"] = 'pagado'
                    else:
                        registros.append(active_res[col])
                        active_res[col] = new_res
                else:
                    active_res[col] = new_res



    for col, res in active_res.items():
        if res:
            registros.append(res)

    return registros, mes, anio, stats


def esc(v):
    if v is None: return "NULL"
    if isinstance(v, (int, float)): return str(v)
    return "'" + str(v).replace("'","''") + "'"

def generate_sql(registros, mes, anio):
    lines = []
    lines.append(f"""-- ================================================================
-- MIGRACIÓN EXCEL → SUPABASE  |  Hotel Verona  |  {mes} {anio}
-- Total reservas a insertar: {len(registros)}
-- ================================================================
-- ⚠️  ANTES DE EJECUTAR:
--   Reemplaza  YOUR_OWNER_ID  con tu UUID real de owner
--   Reemplaza  YOUR_HOTEL_ID  con tu UUID real de hotel
-- ================================================================

BEGIN;

-- ── PASO 1: Huéspedes únicos ──────────────────────────────────
""")
    huespedes = sorted({r["huesped_principal"] for r in registros if r["huesped_principal"]})
    for h in huespedes:
        he = h.replace("'","''")
        correo = re.sub(r'[^a-z0-9]','.', h.lower()) + "@importado.pendiente"
        lines.append(f"""INSERT INTO huespedes (owner_id, nombre_completo, correo)
SELECT 'YOUR_OWNER_ID', '{he}', '{correo}'
WHERE NOT EXISTS (SELECT 1 FROM huespedes WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre_completo)=LOWER('{he}'));
""")

    lines.append("-- ── PASO 2: Empresas únicas ──────────────────────────────────\n")
    empresas = sorted({r["empresa"] for r in registros if r["empresa"]})
    for e in empresas:
        ee = e.replace("'","''")
        lines.append(f"""INSERT INTO empresas (owner_id, nombre, rtn, estado)
SELECT 'YOUR_OWNER_ID', '{ee}', '00000000000000', 'activo'
WHERE NOT EXISTS (SELECT 1 FROM empresas WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre)=LOWER('{ee}'));
""")

    lines.append("-- ── PASO 3: Reservas ─────────────────────────────────────────\n")
    for r in registros:
        ci   = f"{r['fecha']}T15:00:00+00:00" if r['fecha'] else "NULL"
        co_d = r['check_out'] or r['fecha'] or "2026-01-01"
        co_t = "12:00:00" if r['tipo_reserva'] == 'noche' else "19:00:00"
        co   = f"{co_d}T{co_t}+00:00"
        he   = (r["huesped_principal"] or "Sin Nombre").replace("'","''")
        hb   = r["habitacion_nombre"].replace("'","''")
        obs  = (r["observaciones"] or "").replace("'","''")
        emp_join = ""
        emp_sel  = "NULL"
        if r["empresa"]:
            ee = r["empresa"].replace("'","''")
            emp_sel = f"(SELECT id_empresa FROM empresas WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre)=LOWER('{ee}') LIMIT 1)"

        lines.append(f"""INSERT INTO reservas_hotel
    (owner_id,id_huesped,id_hotel,id_habitacion,check_in,check_out,
     adultos,ninos,estado,total_reserva,moneda,estado_pago,
     anticipo,es_cortesia,id_empresa,observaciones,tipo_reserva)
SELECT
    'YOUR_OWNER_ID',
    (SELECT id_huesped FROM huespedes WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre_completo)=LOWER('{he}') LIMIT 1),
    'YOUR_HOTEL_ID',
    (SELECT id_habitacion FROM habitaciones WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre_habitacion)=LOWER('{hb}') LIMIT 1),
    '{ci}','{co}',1,0,
    '{r["estado_reserva"]}',{r["total_reserva"] or 0},'HNL','{r["estado_pago"]}',
    0,{str(r["es_cortesia"]).lower()},{emp_sel},
    '{obs}','{r["tipo_reserva"]}'
WHERE (SELECT id_huesped FROM huespedes WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre_completo)=LOWER('{he}') LIMIT 1) IS NOT NULL
  AND (SELECT id_habitacion FROM habitaciones WHERE owner_id='YOUR_OWNER_ID' AND LOWER(nombre_habitacion)=LOWER('{hb}') LIMIT 1) IS NOT NULL;
""")

    lines.append("COMMIT;\n")
    return "\n".join(lines)


# ── RUN ──────────────────────────────────────────────────────────
path = sys.argv[1] if len(sys.argv) > 1 else "Reservasiones_prueba.xlsx"
registros, mes, anio, stats = extract(path)

print(f"📅 {mes} {anio}  |  {len(registros)} reservas (disponibles omitidas)\n")
print("📊 Desglose:")
for k,v in sorted(stats.items()):
    print(f"   {k:30s}: {v}")

# JSON
json_path = os.path.join(tempfile.gettempdir(), "reservas_final.json")
with open(json_path,"w",encoding="utf-8") as f:
    json.dump(registros, f, ensure_ascii=False, indent=2, default=str)

# SQL
sql = generate_sql(registros, mes, anio)
sql_path = os.path.join(tempfile.gettempdir(), "migracion_final.sql")
with open(sql_path,"w",encoding="utf-8") as f:
    f.write(sql)

print(f"\n✅ {json_path}")
print(f"✅ {sql_path}")

# Muestra representativa
print("\n🔍 Muestra (1 de cada estado_pago):")
visto = set()
for r in registros:
    ep = r["estado_pago"]
    if ep not in visto:
        visto.add(ep)
        print(f"\n  [{ep.upper()}]")
        print(f"   Hab    : {r['habitacion_nombre']}  ({r['fecha']})")
        print(f"   Huésped: {r['huesped_principal']}  |  Empresa: {r['empresa']}")
        print(f"   Precio : L.{r['total_reserva']}  |  Canal: {r['canal_reserva']}")
        print(f"   Raw    : {r['texto_raw'][:80]}")
